from __future__ import annotations

import hashlib
import json
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session

from app.models.import_run import ImportRun, ImportRunItem
from app.models.opportunity import Opportunity
from app.scrapers.emma_scraper import public_solicitations_url
from app.services.score_persistence import score_and_store_opportunity
from app.services.source_service import load_business_profile

EMMA_SOURCE_NAME = "Maryland eMMA"
EMMA_SOURCE_URL = "https://emma.maryland.gov/"
EXCEL_EPOCH = datetime(1899, 12, 30)
REQUIRED_HEADERS = {
    "ID",
    "Title",
    "Status",
    "Due / Close Date",
    "Main Category",
    "Solicitation Type",
    "Issuing Agency",
}
SOURCE_FIELDS = (
    "external_id",
    "title",
    "agency",
    "source_url",
    "opportunity_url",
    "due_date",
    "description_snippet",
    "extraction_confidence",
    "manual_review_needed",
    "source_status",
)


def excel_serial_to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (EXCEL_EPOCH + timedelta(days=float(value))).date()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return (EXCEL_EPOCH + timedelta(days=float(stripped))).date()
        except ValueError:
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(stripped.split()[0], fmt).date()
                except ValueError:
                    continue
    return None


def build_emma_opportunity_url(bpm_id: str) -> str:
    digits = "".join(ch for ch in bpm_id if ch.isdigit())
    opportunity_id = str(int(digits)) if digits else bpm_id
    return f"https://emma.maryland.gov/page.aspx/en/bpm/process_manage_extranet/{opportunity_id}"


def _normalize_header(value: Any) -> str:
    return " ".join(str(value or "").split())


def _description(row: dict[str, Any], publish_date: date | None) -> str:
    parts = [
        f"Status: {row.get('Status')}",
        f"Category: {row.get('Main Category')}",
        f"Type: {row.get('Solicitation Type')}",
    ]
    if publish_date:
        parts.append(f"Published: {publish_date.isoformat()}")
    return " | ".join(part for part in parts if part and not part.endswith(": None"))


def _confidence(row: dict[str, Any]) -> float:
    required_values = [
        row.get("ID"),
        row.get("Title"),
        row.get("Issuing Agency"),
        row.get("Due / Close Date"),
    ]
    present = sum(1 for value in required_values if value not in (None, ""))
    return 0.95 if present == len(required_values) else 0.75


def parse_emma_excel(path: str | Path) -> list[dict]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise ValueError(f"Excel file not found: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("eMMA import requires a .xlsx file")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return _parse_workbook(workbook)
    finally:
        workbook.close()


def parse_emma_excel_bytes(contents: bytes) -> list[dict]:
    if not contents:
        raise ValueError("Excel file is empty")
    workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    try:
        return _parse_workbook(workbook)
    finally:
        workbook.close()


def _parse_workbook(workbook) -> list[dict]:
    sheet = workbook.active
    # eMMA exports can report a stale worksheet dimension of A1, so force
    # openpyxl to stream the actual rows.
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [_normalize_header(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("eMMA workbook is empty") from exc

    missing = REQUIRED_HEADERS.difference(headers)
    if missing:
        raise ValueError(f"eMMA workbook is missing required columns: {sorted(missing)}")

    parsed: list[dict] = []
    for raw_values in rows:
        item = _parse_row(headers, raw_values)
        if item:
            parsed.append(item)
    return parsed


def _parse_row(headers: list[str], raw_values: tuple[Any, ...]) -> dict | None:
    row = dict(zip(headers, raw_values, strict=False))
    status = str(row.get("Status") or "").strip()
    title = str(row.get("Title") or "").strip()
    bpm_id = str(row.get("ID") or "").strip()
    if not title or not bpm_id:
        return None

    due_date = excel_serial_to_date(row.get("Due / Close Date"))
    publish_date = excel_serial_to_date(row.get("Publish Date UTC-4"))
    agency = str(row.get("Issuing Agency") or EMMA_SOURCE_NAME).strip()
    return {
        "external_id": bpm_id,
        "source_status": status,
        "title": f"{bpm_id} {title}"[:500],
        "agency": agency,
        "source_name": EMMA_SOURCE_NAME,
        "source_url": public_solicitations_url(EMMA_SOURCE_URL),
        "opportunity_url": build_emma_opportunity_url(bpm_id),
        "due_date": due_date,
        "description_snippet": _description(row, publish_date),
        "extraction_confidence": _confidence(row),
        "manual_review_needed": False,
    }


def _row_hash(item: dict) -> str:
    payload = {
        field: _serialize_change_value(item.get(field))
        for field in SOURCE_FIELDS
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _find_duplicate(db: Session, item: dict) -> Opportunity | None:
    return (
        db.query(Opportunity)
        .filter(
            or_(
                Opportunity.opportunity_url == item.get("opportunity_url"),
                and_(
                    Opportunity.title == item.get("title"),
                    Opportunity.agency == item.get("agency"),
                    Opportunity.due_date == item.get("due_date"),
                ),
            )
        )
        .first()
    )


def _find_existing(db: Session, item: dict) -> Opportunity | None:
    external_id = item.get("external_id")
    if external_id:
        existing = (
            db.query(Opportunity)
            .filter(
                Opportunity.source_name == item.get("source_name"),
                Opportunity.external_id == external_id,
            )
            .first()
        )
        if existing:
            return existing
    return _find_duplicate(db, item)


def _serialize_change_value(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _source_changes(existing: Opportunity, item: dict) -> dict[str, dict[str, Any]]:
    changes: dict[str, dict[str, Any]] = {}
    for field in SOURCE_FIELDS:
        current = getattr(existing, field)
        incoming = item.get(field)
        if current != incoming:
            changes[field] = {
                "from": _serialize_change_value(current),
                "to": _serialize_change_value(incoming),
            }
    return changes


def _apply_source_updates(existing: Opportunity, item: dict) -> None:
    for field in SOURCE_FIELDS:
        setattr(existing, field, item.get(field))
    existing.last_seen_at = datetime.utcnow()


def _create_opportunity(db: Session, item: dict) -> Opportunity:
    now = datetime.utcnow()
    row = Opportunity(
        title=item["title"],
        agency=item["agency"],
        source_name=item["source_name"],
        source_url=item["source_url"],
        opportunity_url=item.get("opportunity_url"),
        external_id=item.get("external_id"),
        source_status=item.get("source_status"),
        due_date=item.get("due_date"),
        description_snippet=item.get("description_snippet"),
        extraction_confidence=item.get("extraction_confidence", 0.75),
        manual_review_needed=item.get("manual_review_needed", False),
        status="Saved",
        last_seen_at=now,
    )
    db.add(row)
    db.flush()
    return row


def _import_candidate(
    db: Session,
    item: dict,
    profile: dict,
    auto_score: bool,
    counts: dict[str, int],
) -> tuple[str, Opportunity | None, dict | None, str]:
    row_sha = _row_hash(item)
    existing = _find_existing(db, item)
    source_status = str(item.get("source_status") or "").strip().lower()

    if not existing and source_status != "open":
        counts["skipped"] += 1
        return "skipped", None, {"reason": "source status is not open"}, row_sha

    if not existing:
        row = _create_opportunity(db, item)
        counts["created"] += 1
        if auto_score:
            score_and_store_opportunity(db, row, profile)
            counts["scored"] += 1
        return "created", row, None, row_sha

    changes = _source_changes(existing, item)
    existing.last_seen_at = datetime.utcnow()
    if not changes:
        counts["unchanged"] += 1
        return "unchanged", existing, None, row_sha

    _apply_source_updates(existing, item)
    counts["updated"] += 1
    return "updated", existing, changes, row_sha


def _import_run_summary(run: ImportRun) -> dict:
    return {
        "ok": run.status == "completed",
        "import_run_id": run.id,
        "source": run.source_name,
        "filename": run.filename,
        "rows_seen": run.rows_seen,
        "created": run.created,
        "updated": run.updated,
        "unchanged": run.unchanged,
        "skipped": run.skipped,
        "scored": run.scored,
        "mock_fallback_used": False,
    }


def import_emma_excel_upload(
    db: Session,
    contents: bytes,
    filename: str,
    content_type: str | None = None,
    profile: dict | None = None,
    auto_score: bool = True,
) -> dict:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("eMMA import requires a .xlsx file")

    profile = profile if profile is not None else load_business_profile()
    run = ImportRun(
        source_name=EMMA_SOURCE_NAME,
        filename=filename,
        content_type=content_type,
        file_size_bytes=len(contents),
        file_sha256=hashlib.sha256(contents).hexdigest(),
        workbook_bytes=contents,
        status="completed",
    )
    db.add(run)
    db.flush()

    try:
        candidates = parse_emma_excel_bytes(contents)
    except ValueError as exc:
        run.status = "failed"
        run.error_message = str(exc)
        db.commit()
        raise

    counts = {"created": 0, "updated": 0, "unchanged": 0, "skipped": 0, "scored": 0}
    seen_external_ids: set[str] = set()
    for item in candidates:
        external_id = item.get("external_id")
        if external_id and external_id in seen_external_ids:
            action = "skipped"
            opportunity = None
            change_summary = {"reason": "duplicate external_id in workbook"}
            row_sha = _row_hash(item)
            counts["skipped"] += 1
        else:
            if external_id:
                seen_external_ids.add(external_id)
            action, opportunity, change_summary, row_sha = _import_candidate(
                db,
                item,
                profile,
                auto_score,
                counts,
            )
        db.add(
            ImportRunItem(
                import_run=run,
                opportunity=opportunity,
                external_id=item.get("external_id"),
                row_sha256=row_sha,
                action=action,
                change_summary=json.dumps(change_summary, sort_keys=True)
                if change_summary
                else None,
                raw_title=item.get("title"),
                raw_agency=item.get("agency"),
                raw_due_date=item.get("due_date").isoformat() if item.get("due_date") else None,
                raw_source_status=item.get("source_status"),
            )
        )

    run.rows_seen = len(candidates)
    run.created = counts["created"]
    run.updated = counts["updated"]
    run.unchanged = counts["unchanged"]
    run.skipped = counts["skipped"]
    run.scored = counts["scored"]
    db.commit()
    db.refresh(run)
    return _import_run_summary(run)


def import_emma_excel(
    db: Session,
    path: str | Path,
    profile: dict | None = None,
    auto_score: bool = True,
) -> dict:
    profile = profile if profile is not None else load_business_profile()
    candidates = parse_emma_excel(path)
    created = 0
    duplicates_skipped = 0
    scored = 0

    for item in candidates:
        if str(item.get("source_status") or "").strip().lower() != "open":
            continue
        if _find_existing(db, item):
            duplicates_skipped += 1
            continue
        row = _create_opportunity(db, item)
        created += 1
        if auto_score:
            score_and_store_opportunity(db, row, profile)
            scored += 1
    db.commit()

    return {
        "ok": True,
        "source": EMMA_SOURCE_NAME,
        "rows_seen": len(candidates),
        "created": created,
        "duplicates_skipped": duplicates_skipped,
        "scored": scored,
        "mock_fallback_used": False,
    }
